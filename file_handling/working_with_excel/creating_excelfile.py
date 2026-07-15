# run   pip install openpyxl    in terminal once

import openpyxl

# create workbook
wb = openpyxl.Workbook()

# get active worksheet
ws = wb.active
ws.title = "students"

# write to cells
ws["A1"] = "Name"
ws["B1"] = "Grade"
ws["C1"] = "Marks"

ws["A2"] = "Adam"
ws["B2"] = "A"
ws["C2"] = 80

ws["A3"] = "Eve"
ws["B3"] = "B+"
ws["C3"] = 75

# save file
wb.save("students.xlsx")
print("file created")
