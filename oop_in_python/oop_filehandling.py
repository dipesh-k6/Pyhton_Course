import openpyxl, os, sys

class StudentDatabase:
    def __init__(self):
        self.FILE_NAME = "students.xlsx"
        self.WORK_SHEET = "student"

        if not os.path.exists(self.FILE_NAME):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.WORK_SHEET

            ws["A1"] = "Name"
            ws["B1"] = "Address"
            ws["C1"] = "Contact Number"

            wb.save(self.FILE_NAME)

        self.wb = openpyxl.load_workbook(self.FILE_NAME)
        self.ws = self.wb[self.WORK_SHEET]

    def insert_data(self, name, address, contact):
        self.ws.append([name, address, contact])
        self.wb.save(self.FILE_NAME)

    def read_data(self):
        print("\n")
        for row in self.ws.iter_rows(values_only = True):
            print(row)

database = StudentDatabase()

while True:
    choice = None
    valid_choices = ["1", "2", "3"]

    while choice not in valid_choices:
        choice = input("""
        1 to insert data
        2 to read data
        3 to exit
        enter your choice from above: """)

    if choice == "3":
        sys.exit(0)

    elif choice == "2":
        database.read_data()

    elif choice == "1":
        name = input("enter name: ")
        address = input("enter address: ")
        contact = input("enter contact info: ")

        database.insert_data(name, address, contact)
        print("data added!\n")

        
