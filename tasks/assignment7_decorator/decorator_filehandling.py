import openpyxl, os, sys

class StudentData:
    def __init__(self):
        self.FILE_NAME = "students.xlsx"
        self.WORK_SHEET = "student"

        if not os.path.exists(self.FILE_NAME):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.WORK_SHEET

            ws["A1"] = "Name"
            ws["B1"] = "Address"
            ws["C1"] = "Contact Number"

            wb.save(self.FILE_NAME)

        self.wb = openpyxl.load_workbook(self.FILE_NAME)
        self.ws = self.wb[self.WORK_SHEET]

    # method to insert data
    def insert_data(self, name, address, contact):
        self.ws.append([name, address, contact])
        self.wb.save(self.FILE_NAME)

    # method to read data
    def read_data(self):
        print("\n")
        for row in self.ws.iter_rows(values_only = True):
            print(row)

student = StudentData()

#decorartor for authorization
def data_entry(func):
    def wrapper(**kwargs):
        if kwargs.get("username") == "admin" and kwargs.get("password") == "admin@123":
            func(**{"name": kwargs.get("name"), "address": kwargs.get("address"), "contact": kwargs.get("contact")})
            return True
    return wrapper

@data_entry
def data(**kwargs):
    student.insert_data(kwargs.get("name"), kwargs.get("address"), kwargs.get("contact"))


# menu
while True:
    choice = None
    valid_choices = ["1", "2", "3"]

    while choice not in valid_choices:
        choice = input("""
        1 to insert data
        2 to read data
        3 to exit
        enter your choice from above: """)

    if choice == "3":
        sys.exit(0)

    elif choice == "2":
        student.read_data()

    elif choice == "1":
        username = input("enter username: ")
        password = input("enter password: ")

        name = input("enter name: ")
        address = input("enter address: ")
        contact = input("enter contact info: ")
        counter = 4

        while not data(**{"username":username, "password":password, "name":name, "address":address, "contact":contact}):
            counter -= 1
            if counter == 0:
                break
            
            print("incorrect username or password!")
            print(f"you can try again {counter} more times")
            username = input("enter username: ")
            password = input("enter password: ")

        else:
            print("data inserted")

        
