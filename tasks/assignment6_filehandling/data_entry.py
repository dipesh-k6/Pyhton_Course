# data entry software
# enter name, address and contact number of the student

import openpyxl, os, sys

while True:
    valid_choices = [1, 2, 3]
    choice = None
    FILE_NAME = "students.xlsx"

    try:
        while choice not in valid_choices:
            choice = int(input(f"""
            1 to insert data
            2 to read data
            3 to exit
            enter your choice from above options: """))

    except ValueError as e:
        print(f"\nError: Please input correct value!")
        continue

    #to exit the system
    if choice == 3:
        sys.exit(0)

    # to check if file exists and create one if not
    if not  os.path.exists(FILE_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "student"

        ws["A1"] = "Name"
        ws["B1"] = "Address"
        ws["C1"] = "Contact Number"

        wb.save(FILE_NAME)

    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb["student"]
        
    # to add data
    if choice == 1:
        name = input("enter name: ")
        address = input("enter address: ")
        contact = input("enter contact number: ")

        ws.append([name, address, contact])
        wb.save(FILE_NAME)

    #to read data
    elif choice == 2:
        print("\n")
        for rows in ws.iter_rows(values_only=True):
            print(rows)
